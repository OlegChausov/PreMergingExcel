import xlrd3
import openpyxl

print("Напишите путь к вашему файлу склада (копии) по образцу c:\\My_Premerging\\wardrobe_copy.xlsx")
wardrobe = input()
with xlrd3.open_workbook(wardrobe) as workbook: #тут с подкачиваем ключ-название товара: значение-кол-во на складе
    sheet = workbook.sheet_by_index(0)
    vals=[sheet.row_values(rownum) for rownum in range(sheet.nrows)]
    addict={}
    print('step1')
for i in vals: #склад
    if i[1] in range(1, 100000):
        addict[str(i[0])]=int(i[1]) #склад
        
        
print('step2')

print("Напишите путь к файлу с прайсом по образцу c:\\My_Premerging\\pricelist.xlsx")
pricelist = input()

with xlrd3.open_workbook(pricelist) as workbook: #тут открываем файл, который мы будем сверять с первым словарем. Открываем и делаем его списком
    sheet1 = workbook.sheet_by_name('Sheet1')
    vals1=[sheet1.row_values(rownum) for rownum in range(sheet1.nrows)] #прайс
print('step3')
wb1=openpyxl.load_workbook(pricelist)
sheet1=wb1['Sheet1']
for j in vals1: #добавление того, что есть
    if j[0] in addict.keys():
        sheet1.cell(vals1.index(j)+1, column=12).value=addict[j[0]]
print('step4')

print("Напишите куда вам сохранить выполненные файлы по образцу c:\\My_Premerging\\")
folder = input()

wb1.save(folder+"File_New.xlsx")


wb=openpyxl.load_workbook(wardrobe)
sheet = wb['Склад весь']
newsheet=wb.create_sheet('Что не нашлось в прайсе') #Что не нашлось в прайсе но было на складе





temp1=[a[0] for a in vals1[1::]]
temp2=[a for a in vals[1::] if a[0] not in temp1 and a[1] in range(1,10000)] #есть на складе, неет в прайсе

for k in temp2:
    newsheet.append(k)
    

wb.save("c:\\Users\\1amperby\\Desktop\\test_ premerging\\File3.xlsx")

            
    



        

        